#!/usr/bin/env python3
"""
将 Markdown 文档中的对比表转换为 Excel 文件
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 读取 markdown 文件
with open('/Users/chenxiangli/.openclaw/workspace/online-income-methods-comparison.md', 'r', encoding='utf-8') as f:
    content = f.read()

# 创建工作簿
wb = Workbook()
wb.remove(wb.active)  # 删除默认工作表

# 定义样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font = Font(bold=True, size=14)
title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 解析表格的函数
def parse_table(table_text):
    """解析 markdown 表格"""
    lines = table_text.strip().split('\n')
    if len(lines) < 3:
        return None, None
    
    # 解析表头
    header_line = lines[0].strip('|').split('|')
    headers = [h.strip() for h in header_line if h.strip()]
    
    # 跳过分隔符行
    data_lines = lines[2:]
    
    # 解析数据行
    rows = []
    for line in data_lines:
        cells = line.strip('|').split('|')
        row = [c.strip() for c in cells]
        # 过滤空行
        if any(row):
            rows.append(row)
    
    return headers, rows

# 解析章节的函数
def parse_section(section_text):
    """解析章节中的所有表格"""
    # 匹配每个方法（以 #### 开头到下一个 #### 之前）
    pattern = r'#### (.+?)\n\n\|(.+?)\n+(.+?)(?=\n#### |$)'
    matches = re.findall(pattern, section_text, re.DOTALL)
    
    results = []
    for match in matches:
        method_name = match[0].strip()
        table_text = f"|{match[1]}\n{match[2]}"
        headers, rows = parse_table(table_text)
        if headers and rows:
            results.append({
                'name': method_name,
                'headers': headers,
                'rows': rows
            })
    
    return results

# 按章节分割内容
sections = {
    '内容创作与流量变现': r'### 3.1 内容创作与流量变现\n(.+?)(?=### 3.2 |## 4\. |$)',
    '电商类': r'### 3.2 电商类\n(.+?)(?=### 3.3 |## 4\. |$)',
    '技能服务类': r'### 3.3 技能服务类\n(.+?)(?=### 3.4 |## 4\. |$)',
    '投资理财类': r'### 3.4 投资理财类\n(.+?)(?=### 3.5 |## 4\. |$)',
    '租赁与资产变现': r'### 3.5 租赁与资产变现\n(.+?)(?=### 3.6 |## 4\. |$)',
    '薅羊毛与任务类': r'### 3.6 薅羊毛与任务类\n(.+?)(?=### 3.7 |## 4\. |$)',
    '新兴细分领域': r'### 3.7 新兴/细分领域\n(.+?)(?=### 3.8 |## 4\. |$)',
    '企业个人品牌': r'### 3.8 企业/个人品牌\n(.+?)(?=## 4\. |$)',
}

# 解析并创建工作表
for section_name, pattern in sections.items():
    ws = wb.create_sheet(title=section_name)
    
    match = re.search(pattern, content, re.DOTALL)
    if not match:
        print(f"未找到章节: {section_name}")
        continue
    
    section_text = match.group(1)
    
    # 解析各个方法
    # 方法名称格式: 3.1.1 短视频创作 或 3.1.1 短视频创作\n\n|维度|内容|
    method_pattern = r'(?:\d+\.\d+\.\d+ )?(.+?)\n\n(\|.+\|\n)+'
    method_matches = re.findall(method_pattern, section_text, re.DOTALL)
    
    current_row = 1
    
    for method_match in method_matches:
        method_name = method_match[0].strip()
        table_text = method_match[1]
        
        # 写入方法名称
        ws.cell(row=current_row, column=1, value=method_name)
        ws.cell(row=current_row, column=1).font = title_font
        ws.cell(row=current_row, column=1).fill = title_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        current_row += 1
        
        headers, rows = parse_table(table_text)
        if not headers:
            continue
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        current_row += 1
        
        # 写入数据
        for row_data in rows:
            # 确保有足够的列
            while len(row_data) < len(headers):
                row_data.append('')
            
            for col, value in enumerate(row_data[:len(headers)], 1):
                # 处理换行
                value = value.replace('<br>', '\n')
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border
            
            # 设置第一列宽度较窄
            ws.column_dimensions[get_column_letter(1)].width = 20
            ws.column_dimensions[get_column_letter(2)].width = 60
            
            current_row += 1
        
        current_row += 1  # 方法之间的间隔

# 添加快速对比表
ws = wb.create_sheet(title="快速对比总表")

# 读取最后的快速对比表
quick_match = re.search(r'## 附录：快速对比表\n\n(.+?)(?=\n---|\*\*|$)', content, re.DOTALL)
if quick_match:
    table_text = quick_match.group(1)
    headers, rows = parse_table(table_text)
    
    if headers and rows:
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 写入数据
        for row_idx, row_data in enumerate(rows, 2):
            while len(row_data) < len(headers):
                row_data.append('')
            
            for col, value in enumerate(row_data[:len(headers)], 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border
            
            # 设置列宽
            ws.column_dimensions[get_column_letter(1)].width = 18
            ws.column_dimensions[get_column_letter(2)].width = 12
            ws.column_dimensions[get_column_letter(3)].width = 10
            ws.column_dimensions[get_column_letter(4)].width = 10
            ws.column_dimensions[get_column_letter(5)].width = 10
            ws.column_dimensions[get_column_letter(6)].width = 15

# 保存文件
output_path = '/Users/chenxiangli/.openclaw/workspace/online-income-methods-comparison.xlsx'
wb.save(output_path)
print(f"Excel 文件已生成: {output_path}")
